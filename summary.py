import pandas as pd
import numpy as np
from datetime import date
import os
import openpyxl
from openpyxl import load_workbook

today = date.today().strftime('%m %d %Y')
day = date.today().strftime('%d')

# dictionaries of requestors/owners
reqs = {
    "ben.whyte (Ben Whyte)": "CAC",
    "Lisa.Leger <Lisa.Leger@ontario.ca>": "CAC",
    "nicholas.landry (Nicholas Landry)": "CAC",
    "aravinth.ramalingam (Aravinth Ramalingam)": "CAC",
    "lisa.parsons (Lisa Parsons)": "CAC",
    "jingxin.jiang (Jingxin Jiang)": "CAC",
    "colleen.pacione (Colleen Pacione)": "CAC",
    "catherine.ryan (Catherine Ryan)": "CAC",
    "mariaalejandra.gonzalezmoctezuma (Maria Alejandra Gonzalez Moctezuma)": "CAC",
    "cynthia.ogbeide (Cynthia Ogbeide)": "MOF",
    "Tejumade.Adenle@ontario.ca": "MOF",
    "brenda.boyle (Brenda Boyle)": "MOF",
    "bukola.ogeleka (Bukola Ogeleka)": "MOF",
    "tamara.gardner (Tamara Gardner)": "MOF",
    "Sabrina.DiFrancesco@ontario.ca": "MOF",
    "shahadat.hossain (Shahadat Hossain)": "CAC",
    "romnick.galang (Romnick Galang)": "CAC",
    "allana.allen (Allana Allen)": "MOF",
    "Azhar.Ahmad@ontario.ca": "CAC",
    "farzana.aziz (Farzana Aziz)": "MOF",
    "donna.schmitz (Donna Schmitz)": "FFX",
    "jeff.jostes (Jeff Jostes)": "FFX",
    "mohammad.shamsi (Mohammad Shamsi)": "CAC",
    "ayesa.parvin (Ayesa Parvin)": "CAC"
}
owners = {
    "nicholas.landry (Nicholas Landry)": "CAC",
    "ben.whyte (Ben Whyte)": "CAC",
    "lisa.parsons (Lisa Parsons)": "CAC",
    "jingxin.jiang": "CAC",
    "colleen.pacione (Colleen Pacione)": "CAC",
    "catherine.ryan (Catherine Ryan)": "CAC",
    "mariaalejandra.gonzalezmoctezuma (Maria Alejandra Gonzalez Moctezuma)": "CAC",
    "aravinth.ramalingam (Aravinth Ramalingam)": "CAC",
    "bukola.ogeleka (Bukola Ogeleka)": "MOF",
    "tejumade.adenle": "MOF",
    "cynthia.ogbeide (Cynthia Ogbeide)": "MOF",
    "tamara.gardner (Tamara Gardner)": "MOF",
    "sabrina.difrancesco": "MOF",
    "jeff.jostes (Jeff Jostes)": "FFX",
    "mason.graham": "FFX",
    "amit.dudhat": "FFX",
    "francesco.leising (Francesco Leising)": "FFX",
    "fredrick.little (Fredrick Little)": "FFX",
    "andy.herdlein (Andy Herdlein)": "FFX",
    "joe.gero": "FFX",
    "alyssa.dinoto": "FFX",
    "brian.smedley (Brian Smedley)": "FFX",
    "alex.umansky": "FFX",
    "molly.vanryn": "FFX",
    "phil.cannon": "FFX",
    "shahadat.hossain": "CAC",
    "ehizogie.ighile": "FFX",
    "allana.allen (Allana Allen)": "MOF",
    "romnick.galang": "CAC",
    "paul.scott (Paul A Scott)": "MOF",
    "leann.stout": "FFX",
    "chris.bosner": "FFX",
    "brenda.boyle (Brenda Boyle)": "MOF",
    "donna.schmitz (Donna Castello)": "FFX",
    "tyler.cabell": "FFX",
    "joe.gero (Joe Gero)": "FFX",
    "donna.schmitz (Donna Schmitz)": "FFX",
    "kevin.powell (Kevin Powell)": "CAC",
    "farzana.aziz (Farzana Aziz )": "MOF", # extra space for testing
    "mohammad.shamsi (Mohammad Shamsi)": "CAC"
}

# Today RT list of defects
df = pd.read_excel('Results.xlsx').drop(columns=['QueueName', 'Priority', 'Defect #']).replace([np.nan, np.inf, -np.inf], '')
df['CAC/MOF Requestor'] = None
df['Ministry/FFX Owner'] = None
df['CAC/MOF/FFX Owner'] = None

# Note: need to update 'Customer' column to display 5 decimal places instead of 3
# (i.e. show 2.02405E+16 instead of 2.024E+16)

new_defects = 0

for i in range(len(df)):
    df.at[i, 'CAC/MOF Requestor'] = reqs[df.at[i, 'Requestors']]
    df.at[i, 'CAC/MOF/FFX Owner'] = owners[df.at[i, 'OwnerName']]
    df.at[i, 'Ministry/FFX Owner'] = 'FFX' if df.at[i, 'CAC/MOF/FFX Owner'] == 'FFX' else 'Ministry'

# count new defects for the day
for i in reversed(range(len(df))):
    if df.at[i, 'Created'][9:9+len(day)] == '30':
        new_defects += 1
    else:
        break

output_file = today + ' - Defect RT Status Report.xlsx'
writer = pd.ExcelWriter(output_file, engine='xlsxwriter')
df.to_excel(writer, sheet_name='Today RT list of defects', index=False)
writer.close()


# Pivot Table
with pd.ExcelWriter('pivot_table.xlsx', engine='xlsxwriter') as writer:
    p1 = pd.pivot_table(df, values='#', index=['CAC/MOF Requestor'], columns=['CustomField.{Current Status}'], aggfunc='count', fill_value=0)
    p1['Grand Total'] = p1[0:3].sum(axis=1)
    total1 = pd.DataFrame(p1.sum()).T
    total1.index = ['Grand Total']
    p1 = pd.concat([p1, total1]).rename_axis('Row Labels')

    p2 = pd.pivot_table(df, values='#', index=['Status'], columns=['CustomField.{Ticket Severity}'], aggfunc='count', fill_value=0)
    p2['Grand Total'] = p2[0:len(p2)].sum(axis=1)
    total2 = pd.DataFrame(p2.sum()).T
    total2.index = ['Grand Total']
    p2 = pd.concat([p2, total2]).rename_axis('Row Labels')

    p3 = pd.pivot_table(df, values='#', index=['Status'], columns=['CustomField.{Current Status}'], aggfunc='count', fill_value=0)
    p3['Grand Total'] = p3[0:len(p3)].sum(axis=1)
    total3 = pd.DataFrame(p3.sum()).T
    total3.index = ['Grand Total']
    p3 = pd.concat([p3, total3]).rename_axis('Row Labels')

    p4 = pd.pivot_table(df, values='#', index=['CAC/MOF/FFX Owner'], columns=['CustomField.{Ticket Severity}'], aggfunc='count', fill_value=0)
    p4['Grand Total'] = p4[0:len(p4)].sum(axis=1)
    total4 = pd.DataFrame(p4.sum()).T
    total4.index = ['Grand Total']
    p4 = pd.concat([p4, total4]).rename_axis('Row Labels')

    p5 = pd.pivot_table(df, values='#', index=['CustomField.{Ticket Severity}'], columns=['CAC/MOF Requestor'], aggfunc='count', fill_value=0)
    p5['Grand Total'] = p5[0:len(p5)].sum(axis=1)
    total5 = pd.DataFrame(p5.sum()).T
    total5.index = ['Grand Total']
    p5 = pd.concat([p5, total5]).rename_axis('Row Labels')

    p6 = pd.pivot_table(df, values='#', index=['Ministry/FFX Owner'], columns=['CustomField.{Ticket Severity}'], aggfunc='count', fill_value=0)
    p6['Grand Total'] = p6[0:len(p6)].sum(axis=1)
    total6 = pd.DataFrame(p6.sum()).T
    total6.index = ['Grand Total']
    p6 = pd.concat([p6, total6]).rename_axis('Ministry/FFX Owner')

    p7 = pd.pivot_table(df, values='#', index=['CAC/MOF/FFX Owner'], columns=['CustomField.{Ticket Severity}'], aggfunc='count', fill_value=0)
    p7['Grand Total'] = p7[0:len(p7)].sum(axis=1)
    total7 = pd.DataFrame(p7.sum()).T
    total7.index = ['Grand Total']
    p7 = pd.concat([p7, total2]).rename_axis('CAC/MOF/FFX Owner')

    p8 = pd.pivot_table(df, values='#', index=['Status'], columns=['CustomField.{Ticket Severity}'], aggfunc='count', fill_value=0)
    p8['Grand Total'] = p8[0:len(p8)].sum(axis=1)
    total8 = pd.DataFrame(p8.sum()).T
    total8.index = ['Grand Total']
    p8 = pd.concat([p8, total8]).rename_axis('Row Labels')

    p1.to_excel(writer, sheet_name='Pivot Table', startrow=1)

    worksheet = writer.sheets['Pivot Table']

    worksheet.write('A1', 'Count of CAC/MOF Requestor')
    worksheet.write('B1', 'Column Labels')

    p2.to_excel(writer, sheet_name='Pivot Table', startrow=8)
    worksheet.write('A8', 'Count of Status')
    worksheet.write('B8', 'Column Labels')

    p3.to_excel(writer, sheet_name='Pivot Table', startrow=17)
    worksheet.write('A17', 'Count of Status')
    worksheet.write('B17', 'Column Labels')

    p4.to_excel(writer, sheet_name='Pivot Table', startrow=26)
    worksheet.write('A26', 'Count of Status')
    worksheet.write('B26', 'Column Labels')

    p5.to_excel(writer, sheet_name='Pivot Table', startrow=33)
    worksheet.write('A33', 'Count of Status')
    worksheet.write('B33', 'Column Labels')

    p6.to_excel(writer, sheet_name='Pivot Table', startrow=40)
    worksheet.write('A40', 'Count of Ministry/FFX Owner')
    worksheet.write('B40', 'CustomField.{Ticket Severity}')

    p7.to_excel(writer, sheet_name='Pivot Table', startrow=47)
    worksheet.write('A47', 'Count of CAC/MOF/FFX Owner')
    worksheet.write('B47', 'CustomField.{Ticket Severity}')

    p8.to_excel(writer, sheet_name='Pivot Table', startrow=54)
    worksheet.write('A54', 'Count of Status')
    worksheet.write('B54', 'Column Labels')


# Graphs
empty_row = pd.DataFrame([{}])

graphs1 = pd.concat([pd.concat([pd.read_excel('template.xlsx', sheet_name='Graphs', nrows=4, skiprows=3), empty_row], ignore_index=True), empty_row], ignore_index=True).replace([np.nan, np.inf, -np.inf], '').map(lambda x: int(x) if isinstance(x, (int, float)) else x)
graphs1.at[0, 'Number'] = new_defects
#### set defects closed today
graphs1.at[1, 'Number'] = 0
graphs1.at[2, 'Number'] = new_defects + graphs1.at[1, 'Number']

graphs2 = pd.read_excel('pivot_table.xlsx', sheet_name='Pivot Table', nrows=5, skiprows=1).replace([np.nan, np.inf, -np.inf], '')
graphs2 = pd.concat([pd.concat([pd.concat([graphs2, empty_row], ignore_index=True), empty_row], ignore_index=True), empty_row], ignore_index=True).replace([np.nan, np.inf, -np.inf], '').map(lambda x: int(x) if isinstance(x, (int, float)) else x).rename(columns={'Row Labels': 'CAC/MOF Requestor'})

graphs3 = pd.read_excel('pivot_table.xlsx', sheet_name='Pivot Table', nrows=6, skiprows=8).replace([np.nan, np.inf, -np.inf], '')
graphs3 = pd.concat([pd.concat([pd.concat([graphs3, empty_row], ignore_index=True), empty_row], ignore_index=True), empty_row], ignore_index=True).replace([np.nan, np.inf, -np.inf], '').map(lambda x: int(x) if isinstance(x, (int, float)) else x).rename(columns={'Row Labels': 'Status'})

graphs4 = pd.read_excel('pivot_table.xlsx', sheet_name='Pivot Table', nrows=6, skiprows=17).replace([np.nan, np.inf, -np.inf], '')
graphs4 = pd.concat([pd.concat([pd.concat([graphs4, empty_row], ignore_index=True), empty_row], ignore_index=True), empty_row], ignore_index=True).replace([np.nan, np.inf, -np.inf], '').map(lambda x: int(x) if isinstance(x, (int, float)) else x).rename(columns={'Row Labels': 'Row Labels'})

graphs5 = pd.read_excel('pivot_table.xlsx', sheet_name='Pivot Table', nrows=5, skiprows=26).replace([np.nan, np.inf, -np.inf], '')
graphs5 = pd.concat([pd.concat([pd.concat([graphs5, empty_row], ignore_index=True), empty_row], ignore_index=True), empty_row], ignore_index=True).replace([np.nan, np.inf, -np.inf], '').map(lambda x: int(x) if isinstance(x, (int, float)) else x).rename(columns={'Row Labels': 'CAC/MOF/FFX Owner'})

graphs6 = pd.read_excel('pivot_table.xlsx', sheet_name='Pivot Table', nrows=5, skiprows=33).replace([np.nan, np.inf, -np.inf], '')
graphs6 = pd.concat([pd.concat([pd.concat([graphs6, empty_row], ignore_index=True), empty_row], ignore_index=True), empty_row], ignore_index=True).replace([np.nan, np.inf, -np.inf], '').map(lambda x: int(x) if isinstance(x, (int, float)) else x).rename(columns={'Row Labels': 'Severity by Requestor'})

graphs7 = pd.read_excel('pivot_table.xlsx', sheet_name='Pivot Table', nrows=4, skiprows=40).replace([np.nan, np.inf, -np.inf], '').drop(columns=['Severity 2', 'Severity 3', 'Severity 4']).rename(columns={'Grand Total': 'Number of Tickets'})
graphs7 = pd.concat([pd.concat([pd.concat([graphs7, empty_row], ignore_index=True), empty_row], ignore_index=True), empty_row], ignore_index=True).replace([np.nan, np.inf, -np.inf], '').map(lambda x: int(x) if isinstance(x, (int, float)) else x).rename(columns={'Ministry/FFX Owner': 'Ticket Owner'})

graphs8 = pd.read_excel('pivot_table.xlsx', sheet_name='Pivot Table', nrows=5, skiprows=47).replace([np.nan, np.inf, -np.inf], '').drop(columns=['Grand Total', 'Severity 3', 'Severity 4']).rename(columns={'Severity 2': '# Sev 2 Tickets'})
graphs8 = pd.concat([pd.concat([pd.concat([graphs8, empty_row], ignore_index=True), empty_row], ignore_index=True), empty_row], ignore_index=True).replace([np.nan, np.inf, -np.inf], '').map(lambda x: int(x) if isinstance(x, (int, float)) else x).rename(columns={'CAC/MOF/FFX Owner': 'Ticket Owner'})

graphs9 = pd.read_excel('pivot_table.xlsx', sheet_name='Pivot Table', nrows=5, skiprows=47).replace([np.nan, np.inf, -np.inf], '').drop(columns=['Severity 2', 'Severity 3', 'Severity 4']).rename(columns={'Grand Total': 'Total # Tickets'})
graphs9 = pd.concat([pd.concat([pd.concat([graphs9, empty_row], ignore_index=True), empty_row], ignore_index=True), empty_row], ignore_index=True).replace([np.nan, np.inf, -np.inf], '').map(lambda x: int(x) if isinstance(x, (int, float)) else x).rename(columns={'CAC/MOF/FFX Owner': 'Ticket Owner'})

with pd.ExcelWriter('combined_tables.xlsx', engine='xlsxwriter') as writer:
    graphs1.to_excel(writer, sheet_name='Sheet1', startrow=0, index=False)
    graphs2.to_excel(writer, sheet_name='Sheet1', startrow=6, index=False)
    graphs3.to_excel(writer, sheet_name='Sheet1', startrow=14, index=False)
    graphs4.to_excel(writer, sheet_name='Sheet1', startrow=23, index=False)
    graphs5.to_excel(writer, sheet_name='Sheet1', startrow=33, index=False)
    graphs6.to_excel(writer, sheet_name='Sheet1', startrow=41, index=False)
    graphs7.to_excel(writer, sheet_name='Sheet1', startrow=48, index=False)
    graphs8.to_excel(writer, sheet_name='Sheet1', startrow=55, index=False)
    graphs9.to_excel(writer, sheet_name='Sheet1', startrow=62, index=False)


# Merging files
file_list = [output_file, 'pivot_table.xlsx', 'combined_tables.xlsx']

combined_workbook = openpyxl.Workbook()
combined_workbook.remove(combined_workbook.active)

for filename in file_list:
    file_path = os.path.join(os.getcwd(), filename)
    workbook = load_workbook(file_path)
    
    for sheet_name in workbook.sheetnames:
        source_sheet = workbook[sheet_name]
        combined_sheet = combined_workbook.create_sheet(title=f"{os.path.splitext(filename)[0]}_{sheet_name}")

        for row in source_sheet.iter_rows():
            for cell in row:
                combined_sheet[cell.coordinate].value = cell.value
                combined_sheet[cell.coordinate].number_format = cell.number_format

combined_workbook.save(output_file)

# New writer for styling
writer = pd.ExcelWriter(output_file, engine='xlsxwriter')

# Styling "Today RT list of defects" spreadsheet
df.to_excel(writer, sheet_name='Today RT list of defects', index=False)
workbook = writer.book
worksheet = writer.sheets['Today RT list of defects']

header_format = workbook.add_format({
    'bold': True,
    'valign': 'top',
    'fg_color': '#4472c4',
    'font_color': "#ffffff",
    'border': 1,
    'border_color': '#8ea9db' 
})

odd_format = workbook.add_format({
    'valign': 'top',
    'fg_color': '#d9e1f2',
    'border': 1,
    'border_color': '#8ea9db' 
})

for col_num, value in enumerate(df.columns.values):
    worksheet.write(0, col_num, value, header_format)

for row_num, row_data in df.iterrows():
    if (row_num + 1) % 2 != 0:
        for col_num, value in enumerate(df.columns.values):
            worksheet.write(row_num + 1, col_num, row_data[value], odd_format)

worksheet.set_column('A:O', 10)

# styling Pivot Table spreadsheet
pivot_list = [p1, p2, p3, p4, p5, p6, p7, p8]
headers = [1, 7, 8, 16, 17, 25, 26, 32, 33, 39, 40, 43, 46, 47, 53, 54]
if(len(p1) == 2):
    headers.append(3)
elif(len(p1) == 3):
    headers.append(4)
elif(len(p1) == 4):
    headers.append(5)
if(len(p2) == 2):
    headers.append(10)
elif(len(p2) == 3):
    headers.append(11)
elif(len(p2) == 4):
    headers.append(12)
elif(len(p2) == 5):
    headers.append(13)
if(len(p3) == 2):
    headers.append(19)
elif(len(p3) == 3):
    headers.append(20)
elif(len(p3) == 4):
    headers.append(21)
elif(len(p3) == 5):
    headers.append(22)
if(len(p4) == 2):
    headers.append(28)
elif(len(p4) == 3):
    headers.append(29)
elif(len(p4) == 4):
    headers.append(30)
if(len(p5) == 2):
    headers.append(35)
elif(len(p5) == 3):
    headers.append(36)
elif(len(p5) == 4):
    headers.append(37)
if(len(p7) == 2):
    headers.append(49)
elif(len(p7) == 3):
    headers.append(50)
elif(len(p7) == 4):
    headers.append(51)
if(len(p8) == 2):
    headers.append(56)
elif(len(p8) == 3):
    headers.append(57)
elif(len(p8) == 4):
    headers.append(58)
elif(len(p8) == 5):
    headers.append(59)
headers.sort()

pivots = pd.read_excel('pivot_table.xlsx').replace([np.nan, np.inf, -np.inf], '')

pivots.to_excel(writer, sheet_name='Pivot Table', index=False)

workbook = writer.book
worksheet = writer.sheets['Pivot Table']

blank_format = workbook.add_format({})

header_format = workbook.add_format({
    'fg_color': '#d9e1f2',
    'border': 1,
    'border_color': '#8ea9db'
})

# styling first header
for col_num, value in enumerate(pivots.columns.values):
    if(value[0:8] == 'Unnamed:'):
        if(col_num <= pivot_list[0].shape[1]):
            worksheet.write(0, col_num, '', header_format)
        else:
            worksheet.write(0, col_num, '', blank_format)
    else:
        worksheet.write(0, col_num, value, header_format)

# styling remaining headers
pivot_list_index = 0
for i in range(len(headers)):
    if(abs(i - 2) % 3 == 0):
        pivot_list_index += 1
    for col_num, value in enumerate(pivots.iloc[headers[i] - 1]):
        if(value == '' or (len(str(value)) >= 10 and str(value)[0:8] == 'Unnamed:')):
            if(col_num <= pivot_list[pivot_list_index].shape[1]):
                worksheet.write(headers[i], col_num, '', header_format)
            else:
                worksheet.write(headers[i], col_num, '', blank_format)
        else:
            worksheet.write(headers[i], col_num, value, header_format)

worksheet.set_column('A:Z', 16)

# styling Graphs spreadsheet
headers = [3, 6, 6+len(graphs2)-3, 14, 14+len(graphs3)-3, 23, 23+len(graphs4)-3, 33, 33+len(graphs5)-3, 41, 41+len(graphs6)-3, 48, 51, 55, 59, 62, 66]
shaded = [2, 8, 16, 25, 35, 43, 50, 57, 64]
if(len(graphs3) >= 8):
    shaded.append(18)
if(len(graphs4) >= 8):
    shaded.append(27)
not_shaded = [1, 7, 15, 17, 24, 26, 34, 42, 44, 49, 56, 63]
if(len(graphs2) >= 7):
    not_shaded.append(9)
if(len(graphs5) >= 7):
    not_shaded.append(36)
if(len(graphs8) >= 7):
    not_shaded.append(58)
if(len(graphs9) >= 7):
    not_shaded.append(65)

graphs = pd.read_excel('combined_tables.xlsx').replace([np.nan, np.inf, -np.inf], '')
graphs.to_excel(writer, sheet_name='Graphs', index=False)

workbook = writer.book
worksheet = writer.sheets['Graphs']

blank_format = workbook.add_format({})

header_format = workbook.add_format({
    'bold': True,
    'valign': 'top',
    'fg_color': '#4472c4',
    'font_color': "#ffffff",
    'border': 2,
    'border_color': '#000000',
    'align': 'center'
})

shaded_format = workbook.add_format({
    'valign': 'top',
    'fg_color': '#d9e1f2',
    'border': 2,
    'border_color': '#000000',
    'align': 'center'
})

not_shaded_format = workbook.add_format({
    'valign': 'top',
    'border': 2,
    'border_color': '#000000',
    'align': 'center'
})

# styling first header
for col_num, value in enumerate(graphs.columns.values):
    if(value[0:8] == 'Unnamed:'):
        worksheet.write(0, col_num, '', blank_format)
    else:
        worksheet.write(0, col_num, value, header_format)

# styling remaining headers
for row_num in headers:
    for col_num, value in enumerate(graphs.iloc[row_num - 1]):
        if(value == '' or (len(str(value)) >= 10 and str(value)[0:8] == 'Unnamed:')):
            worksheet.write(row_num, col_num, '', blank_format)
        else:
            worksheet.write(row_num, col_num, value, header_format)

for row_num in shaded:
    for col_num, value in enumerate(graphs.iloc[row_num - 1]):
        if(value != ''):
            worksheet.write(row_num, col_num, value, shaded_format)

for row_num in not_shaded:
    for col_num, value in enumerate(graphs.iloc[row_num - 1]):
        if(value != ''):
            worksheet.write(row_num, col_num, value, not_shaded_format)

worksheet.set_column('A:Z', 16)

# Inserting graphs
bar_graph = workbook.add_chart({'type': 'column', 'subtype': 'stacked'})
bar_graph.add_series({
    'name': '=Graphs!$B$34',
    'categories': '=Graphs!$A$35:$A$38',
    'values': '=Graphs!$B$35:$B$38',
    'fill': {'color': '#5B9BD5'}
})
bar_graph.add_series({
    'name': '=Graphs!$C$34',
    'categories': '=Graphs!$A$35:$A$38',
    'values': '=Graphs!$C$35:$C$38',
    'fill': {'color': '#ED7D31'}
})
bar_graph.add_series({
    'name': '=Graphs!$D$34',
    'categories': '=Graphs!$A$35:$A$38',
    'values': '=Graphs!$D$35:$D$38',
    'fill': {'color': '#A5A5A5'}
})
bar_graph.set_title({
    'name': 'Defect Status\nSeverity by Owner as of ' + date.today().strftime('%B %-d'),
    'alignment': {
        'horizontal': 'center'
    },
    'name_font': {
        'bold': False,
        'size': 16
    },
    'points': [
        {'fill': {'color': '#5B9BD5'}},
        {'fill': {'color': '#ED7D31'}},
        {'fill': {'color': '#A5A5A5'}}
    ]
})
bar_graph.set_y_axis({
    'major_unit': 2
})
bar_graph.set_table({
    'show_keys': True
})
bar_graph.set_legend({'position': 'bottom'})
worksheet.insert_chart('G34', bar_graph)

tickets_chart = workbook.add_chart({'type': 'pie'})
tickets_chart.add_series({
    'categories': ['Graphs', 49, 0, 50, 0],  # Categories (x-axis)
    'values':     ['Graphs', 49, 1, 50, 1],  # Values (y-axis)
    'data_labels': {'value': True},
    'points': [
        {'fill': {'color': '#5B9BD5'}},
        {'fill': {'color': '#ED7D31'}},
    ]
})
tickets_chart.set_title({
    'name': 'TICKETS BY OWNER: ' + str(graphs9.iloc[3, 1]),
    'alignment': {
        'horizontal': 'center'
    },
    'name_font': {
        'bold': False,
        'size': 12
    },
})
tickets_chart.set_legend({'position': 'bottom'})
worksheet.insert_chart('C49', tickets_chart, {
    'x_offset': 25,
    'y_offset': 10,
    'x_scale': 0.5,
    'y_scale': 1.0
})

sev2_chart = workbook.add_chart({'type': 'pie'})
sev2_chart.add_series({
    'categories': ['Graphs', 56, 0, 58, 0],  # Categories (x-axis)
    'values':     ['Graphs', 56, 1, 58, 1],  # Values (y-axis)
    'data_labels': {'value': True},
    'points': [
        {'fill': {'color': '#5B9BD5'}},
        {'fill': {'color': '#ED7D31'}},
        {'fill': {'color': '#A5A5A5'}}
    ]
})
sev2_chart.set_title({
    'name': 'Total Number of sev 2 Tickets by owner : ' + str(graphs8.iloc[3, 1]),
    'alignment': {
        'horizontal': 'center'
    },
    'name_font': {
        'bold': False,
        'size': 12
    },
})
sev2_chart.set_legend({'position': 'bottom'})
worksheet.insert_chart('C64', sev2_chart, {
    'x_offset': 25,
    'y_offset': 10,
    'x_scale': 0.5,
    'y_scale': 1.0
})

# Styling complete
writer.close()
os.remove("pivot_table.xlsx")
os.remove("combined_tables.xlsx")



# formatting bug: since lists used for formatting are fixed and 
# number of rows can vary from day to day, sometimes the formats
# will be off -> see June 4 defect report